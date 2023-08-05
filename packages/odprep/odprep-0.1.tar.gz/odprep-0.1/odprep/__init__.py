# imports
from gbdxtools import Interface
gbdx = Interface()
import openpyxl
import pprint
pp = pprint.PrettyPrinter(indent=4)
import time
import datetime
import requests
import json

bucket_location = 's3://disaster-open-data'
xl_in = raw_input("path to excel file: ")
print "user input: ", xl_in


def data_from_xl(xl_path):
    input_xl = openpyxl.load_workbook(xl_path)
    input_sheet = input_xl.get_sheet_by_name('Sheet1')
    event_name = input_sheet['A2'].value
    event_date = input_sheet['B2'].value
    catalog_ids = []
    for i in range(2, input_sheet.max_row + 1):
        cat_id = input_sheet.cell(row=i, column=3)
        catalog_ids.append(cat_id.value)
    return event_name, event_date, catalog_ids


def order_images(catalog_ids):
    order_id = gbdx.ordering.order(catalog_ids)
    print "order ID: " + order_id
    order_info = gbdx.ordering.status(order_id)
    print 'order info: '
    pp.pprint(order_info)
    ready_list = []
    later_list = []
    for i in range(len(order_info)):
        if order_info[i]['state'] == 'delivered':
            ready = order_info[i]['acquisition_id']
            ready_list.append(ready)
        else:
            not_ready = order_info[i]['acquisition_id']
            later_list.append(not_ready)
    print 'Delivered Catalog IDs: '
    pp.pprint(ready_list)
    print 'Submitted Catalog IDs: '
    pp.pprint(later_list)
    status_list = []
    while len(status_list) < len(order_info):
        for i in range(len(order_info)):
            if order_info[i]['state'] == 'delivered' and order_info[i]['acquisition_id'] not in status_list:
                status_list.append(order_info[i]['acquisition_id'])
            else:
                time.sleep(600)
                now = datetime.datetime.now()
                print now.strftime('Order status at %Y/%m/%d %H:%M:')
                print order_info
    return order_info


def get_acquisition_date(catalog_id):
    token = "Bearer %s" % gbdx.gbdx_connection.access_token
    url = 'https://geobigdata.io/catalog/v1/record/' + str(catalog_id) + '?includeRelationships=false'
    headers = {'Content-Type': 'application/json',"Authorization": token}
    catID_result = requests.get(url, headers=headers, data=json.dumps(catalog_id))
    acquisition_date = catID_result.json()['properties']['timestamp']
    return acquisition_date


# ingest event information through csv
info = data_from_xl(xl_in)
name = info[0]
date = info[1]
cat_ids = info[2]
pp.pprint(info)


# order catalog ids
order = order_images(cat_ids)

# process catalog ids, rename and copy to open data bucket
for i in range(len(order)):
    cat_id = order[i]['acquisition_id']
    s3_source = order[i]['location']
    acq_date = get_acquisition_date(cat_id)
    if cat_id[2] == '2':
        aop_task = gbdx.Task("AOP_Strip_Processor", data=s3_source, enable_acomp=False, enable_pansharpen=False,
                             enable_dra=False, ortho_tiling_scheme="DGHalfMeter:18", enable_tiling=True)
    else:
        aop_task = gbdx.Task("AOP_Strip_Processor", data=s3_source, ortho_tiling_scheme="DGHalfMeter:18",
                             enable_tiling=True, dra_mode="BaseLayerMatch")
    prep_task = gbdx.Task("open_data_prep", input_directory=aop_task.outputs.data.value, catalog_id=cat_id,
                          event_name=name, event_date=date, acquisition_date=acq_date)
    s3_task = gbdx.Task("StageDataToS3", data=prep_task.outputs.output_directory.value, destination=bucket_location)
    workflow = gbdx.Workflow([ aop_task, prep_task, s3_task ])
    workflow.execute()
    print 'Catalog ID: ', cat_id, ',', 'Workflow ID: ', workflow.id

















# return something and print to csv?


# def data_from_csv(csv_path):
#     input_csv = open(csv_path)
#     input_reader = csv.reader(input_csv)
#     input_list = list(input_reader)
#     event_name = input_list[1][0]
#     event_date = input_list[1][1]
#     catalog_ids = []
#     for i in range(1, (len(input_list))):
#         get_cat = input_list[i][2]
#         catalog_ids.append(get_cat)
#     return event_name, event_date, catalog_ids

# # process catalog ids, rename and copy to open data bucket
# for i in range(len(order)):
#     cat_id = order[i]['acquisition_id']
#     s3_source = order[i]['location']
#     acq_date = get_acquisition_date(cat_id)
#     # stage_bucket = 's3://disaster-open-data/stage/'
#     if cat_id[2] == '2':
#         aop_task = gbdx.Task("AOP_Strip_Processor", data=s3_source, enable_acomp=False, enable_pansharpen=False,
#                              enable_dra=False, ortho_tiling_scheme="DGHalfMeter:18", enable_tiling=True)
#     else:
#         aop_task = gbdx.Task("AOP_Strip_Processor", data=s3_source, ortho_tiling_scheme="DGHalfMeter:18",
#                              enable_tiling=True, dra_mode="BaseLayerMatch")
#     # aop_s3_task = gbdx.Task("StageDataToS3", data=aop_task.outputs.data.value, destination=stage_bucket)
#     prep_task = gbdx.Task("open_data_prep", input_directory=aop_task.outputs.data.value, catalog_id=cat_id,
#                           event_name=name, event_date=date, acquisition_date=acq_date)
#     # prep_task = gbdx.Task("open_data_prep", input_directory=stage_bucket, catalog_id=cat_id,
#     #                       event_name=name, event_date=date, acquisition_date=acq_date)
#     s3_task = gbdx.Task("StageDataToS3", data=prep_task.outputs.output_directory.value, destination=bucket_location)
#     # workflow = gbdx.Workflow([ aop_task, aop_s3_task ])
#     workflow = gbdx.Workflow([ aop_task, prep_task, s3_task ])
#     # workflow = gbdx.Workflow([ prep_task, s3_task ])
#     workflow.execute()
#     print workflow.id