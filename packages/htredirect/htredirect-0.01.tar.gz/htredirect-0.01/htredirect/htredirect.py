# -*- coding: utf-8 -*-
import openpyxl

class HtRedirect:
    def __init__(self, filename):
        wb = openpyxl.load_workbook(filename=filename)
        self.ws = wb.active
        self.redirects = ''

    def generate(self):
        for row in range(self.ws.max_row):
            src = self.ws['{}{}'.format('A', row+1)].value
            dest = self.ws['{}{}'.format('B', row+1)].value
            redirect_type = self.ws['{}{}'.format('C', row+1)].value
            self.redirects += 'RewriteRule ^{} {}$1 [R={},L]\n'.format(src, dest, redirect_type)

    def save(self, filename):
        f = open(filename, 'w')
        f.write(self.redirects)
        f.close()
