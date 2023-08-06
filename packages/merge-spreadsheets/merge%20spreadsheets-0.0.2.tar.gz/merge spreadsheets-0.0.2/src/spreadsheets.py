import copy
from collections import OrderedDict

import openpyxl


class Table:

    def __init__(self, headers=None, rows=None):
        if rows is None:
            rows = []

        self.headers = headers
        self.rows = rows

    def _new_row(self):
        return [None] * len(self.headers)

    def add(self, row):
        new_row = self._new_row()

        for i, k in enumerate(row.keys()):
            if k in self.headers:
                new_row[i] = row[k]
            else:
                raise Exception("Column '{}' does not exists".format(k))

        self.rows.append(new_row)

    @property
    def rows_count(self):
        return len(self.rows)

    @property
    def columns_count(self):
        return len(self.headers)

    @classmethod
    def from_xlsx(cls, filename):
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb.worksheets[0]

        headers = []
        rows = []

        c = 1
        while True:
            cell = ws.cell(row=1, column=c)

            if cell.value is None:
                break
            c = c + 1

            headers.append(cell.value)

        columns_count = len(headers)

        r = 2
        while True:
            row = [
                ws.cell(row=r, column=c).value for
                c in range(1, columns_count+1)
            ]

            if row.count(None) == columns_count:
                break

            rows.append(row)
            r = r + 1

        return cls(headers=tuple(headers), rows=rows)

    def merge(self, other):
        if set(self.headers) != set(other.headers):
            # todo: check duplicates column names
            raise Exception('Tables with differences columns can not be merged')
        pass

        row_template = OrderedDict.fromkeys(self.headers)
        for r in other.rows:
            row = copy.copy(row_template)

            for i, c in enumerate(other.headers):
                row[c] = r[i]

            # todo: check headers
            self.rows.append(row.values())

    def print(self):
        print('|'.join(self.headers))
        for r in self.rows:
            print('|'.join(str(i) for i in r))

    def save_to_file(self, filename, template=None):
        work_book = openpyxl.Workbook()
        sheet = work_book.worksheets[0]
        for i, h in enumerate(self.headers, start=1):
            sheet.cell(row=1, column=i).value = h

        for r, row in enumerate(self.rows, start=2):
            for c, value in enumerate(row, start=1):
                sheet.cell(row=r, column=c).value = value

        work_book.save(filename)


def merge_tables(tables):
    merged_table = Table(headers=tables[0].headers)

    for t in tables:
        merged_table.merge(t)

    return merged_table


def merge_files(files, output_filename):
    tables = [Table.from_xlsx(f) for f in files]
    merged_table = merge_tables(tables)

    merged_table.save_to_file(output_filename)

def merge_directory(directory, result_filename):
    import os
    files = (
        os.path.join(directory, f) for f in os.listdir(directory)
        if (
            f.endswith('.xlsx') and
            f != result_filename
        )
    )

    merge_files(files, os.path.join(directory, result_filename))

